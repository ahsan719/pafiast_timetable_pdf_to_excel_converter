import pdfplumber
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- CONFIGURATION: TIME SLOTS ---
SLOT_MAP_START = {
    1: "09:00", 2: "09:30", 3: "10:00", 4: "10:30",
    5: "11:00", 6: "11:30", 7: "12:00", 8: "12:30",
    9: "13:00", 10: "13:30", 11: "14:00", 12: "14:30",
    13: "15:00", 14: "15:30", 15: "16:00", 16: "16:30"
}

SLOT_MAP_END = {
    1: "09:30", 2: "10:00", 3: "10:30", 4: "11:00",
    5: "11:30", 6: "12:00", 7: "12:30", 8: "13:00",
    9: "13:30", 10: "14:00", 11: "14:30", 12: "15:00",
    13: "15:30", 14: "16:00", 15: "16:30", 16: "17:00"
}

def get_slot_number_from_header(header_text):
    """Extracts the slot number (1-16) from a header cell."""
    if not header_text: return None
    match = re.search(r'^(\d+)', str(header_text).strip())
    if match:
        num = int(match.group(1))
        if 1 <= num <= 16: return num
    return None

def process_pdf_to_data(pdf_path):
    """
    Extracts class data from PDF.
    Returns a list of dictionaries.
    """
    extracted_data = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # 1. Identify Room and Section from page text
            page_text = page.extract_text() or ""
            lines = [line.strip() for line in page_text.split('\n') if line.strip()]
            
            # Extract Room
            # Strategy: Look for pattern in first few lines, or trust the first line if short.
            # Regex covers: A1-109, C1-B05, B2-202, Bio Lab, etc.
            room_name = "Unknown Room"
            room_regex = r'\b([A-Z]\d-[A-Z0-9\-]+|[A-Za-z\s]+(?:Lab|Hall|Auditorium|Room))\b'
            
            # Prioritize searching in the first 3 lines
            header_text = "\n".join(lines[:3])
            room_match = re.search(room_regex, header_text)
            
            if room_match:
                room_name = room_match.group(1)
            elif lines and len(lines[0]) < 15 and re.match(r'^[A-Z0-9\-]+$', lines[0]):
                 # Fallback: First line is short and alphanumeric (e.g., "A1-109")
                room_name = lines[0]
            
            # Extract Section (e.g., BSAI-F24 Green)
            section_match = re.search(r'(BS[A-Z]+-(?:F|S)\d+\s+(?:Green|Blue|Red|White|[A-Z]))', page_text)
            section_name = section_match.group(1) if section_match else "Unknown Section"

            # 2. Extract Table
            table_settings = {
                "vertical_strategy": "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance": 3,
                "join_tolerance": 3,
            }
            table = page.extract_table(table_settings)
            if not table:
                table = page.extract_table()
                if not table: continue

            # 3. Find Header
            col_to_slot = {}
            header_row_idx = -1
            
            for row_idx, row in enumerate(table):
                row_str = " ".join([str(c) if c else "" for c in row])
                if re.search(r'\b1\b', row_str) and re.search(r'\b2\b', row_str):
                    header_row_idx = row_idx
                    for col_idx, cell in enumerate(row):
                        slot_num = get_slot_number_from_header(cell)
                        if slot_num:
                            col_to_slot[col_idx] = slot_num
                    break
            
            if header_row_idx == -1: continue

            # 4. Process Data Rows
            for row_idx in range(header_row_idx + 1, len(table)):
                row = table[row_idx]
                if not row or len(row) == 0: continue
                
                day_cell = (row[0] or "").strip().replace('\n', '')
                if day_cell not in ['Mo', 'Tu', 'We', 'Th', 'Fr']: continue
                
                col_idx = 1
                while col_idx < len(row):
                    cell_content = row[col_idx]
                    
                    if not cell_content or str(cell_content).strip() == "":
                        col_idx += 1
                        continue
                    
                    start_slot = col_to_slot.get(col_idx)
                    if not start_slot:
                        col_idx += 1
                        continue
                    
                    # Span Detection
                    span_count = 1
                    next_col = col_idx + 1
                    while next_col < len(row):
                        next_cell = row[next_col]
                        if next_cell is None and next_col in col_to_slot:
                            span_count += 1
                            next_col += 1
                        else:
                            break
                    
                    end_slot = start_slot + span_count - 1
                    if end_slot > 16: end_slot = 16
                    
                    start_time = SLOT_MAP_START.get(start_slot, "Unknown")
                    end_time = SLOT_MAP_END.get(end_slot, "Unknown")
                    
                    # Process Content
                    raw_text = str(cell_content)
                    class_parts = re.split(r'\n\s*\n', raw_text.strip())
                    
                    for part in class_parts:
                        clean_info = part.replace('\n', ' ').strip()
                        clean_info = re.sub(r'\s+', ' ', clean_info)
                        clean_info = re.sub(r'(BS[A-Z]+-F\d+\s+(?:Green|Blue|Red|White)|Mo|Tu|We|Th|Fr)\b', '', clean_info).strip()
                        
                        if len(clean_info) < 5: continue
                        
                        extracted_data.append({
                            "Room": room_name,
                            "Section": section_name,
                            "Day": day_cell,
                            "Start Slot": start_slot, # Helpful for sorting
                            "Start Time": start_time,
                            "End Time": end_time,
                            "Class Info": clean_info
                        })
                    
                    col_idx = next_col
                    
    return extracted_data

def generate_formatted_excel(data, output_path):
    """
    Generates Excel with valid formatting:
    - Groups by Room.
    - Merged Highlighted Header for each Room.
    """
    if not data: return False
    
    # Sort: Room -> Day Order -> Start Time
    df = pd.DataFrame(data)
    
    # Custom sort for Day
    day_order = {'Mo': 0, 'Tu': 1, 'We': 2, 'Th': 3, 'Fr': 4}
    df['DayOrder'] = df['Day'].map(day_order)
    df = df.sort_values(by=['Room', 'DayOrder', 'Start Slot'])
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Timetable"
    
    # Styles
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Blue
    header_font = Font(color="FFFFFF", bold=True, size=14)
    sub_header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border_style = Side(style='thin')
    thin_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    # Columns
    columns = ["Day", "Start Time", "End Time", "Class Info"]
    
    # Group by Room
    rooms = df['Room'].unique()
    
    current_row = 1
    
    for room in rooms:
        # Get room data
        room_data = df[df['Room'] == room]
        
        # 1. Merged Room Header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(columns))
        cell = ws.cell(row=current_row, column=1)
        cell.value = f"Room: {room}"  # Highlighted Room Name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        
        current_row += 1
        
        # 2. Sub-headers
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = col_name
            cell.font = sub_header_font
            cell.alignment = center_align
            cell.border = thin_border
            
            # Set column width
            if col_name == "Class Info":
                ws.column_dimensions[chr(64+col_idx)].width = 50
            else:
                ws.column_dimensions[chr(64+col_idx)].width = 15
        
        current_row += 1
        
        # 3. Data Rows
        for _, row_data in room_data.iterrows():
            # Day
            cell = ws.cell(row=current_row, column=1)
            cell.value = row_data['Day']
            cell.alignment = center_align
            cell.border = thin_border
            
            # Start
            cell = ws.cell(row=current_row, column=2)
            cell.value = row_data['Start Time']
            cell.alignment = center_align
            cell.border = thin_border
            
            # End
            cell = ws.cell(row=current_row, column=3)
            cell.value = row_data['End Time']
            cell.alignment = center_align
            cell.border = thin_border
            
            # Info
            cell = ws.cell(row=current_row, column=4)
            cell.value = row_data['Class Info']
            cell.alignment = left_align
            cell.border = thin_border
            
            current_row += 1
            
        # Add spacing between classrooms
        current_row += 2
        
    wb.save(output_path)
    return True
